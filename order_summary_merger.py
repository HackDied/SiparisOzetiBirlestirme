#!/usr/bin/env python3
"""
Sipariş Özeti Birleştirme Aracı
Sipariş özeti Excel dosyalarını birleştir (alış fiyatları dahil)
"""

import customtkinter as ctk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path
import pandas as pd
import threading
import sys
import re
import os
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import time
import urllib.request
import urllib.error

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    HAS_DND = True
except ImportError:
    HAS_DND = False

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

CURRENCY_SYMBOLS = {
    'EUR': '€', 'USD': '$', 'TRY': '₺',
}


def _get_script_dir():
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent
    return Path(__file__).parent


SETTINGS_FILE = _get_script_dir() / '.order_merger_settings.json'


COST_CURRENCY_MAP = {
    'TL': 'TRY', 'TRY': 'TRY', '₺': 'TRY',
    'USD': 'USD', '$': 'USD',
    'EUR': 'EUR', '€': 'EUR',
}


def _parse_cost(value):
    """COST string'ini (tutar, para_birimi) olarak çevir: '21500.00 TL' -> (21500.0, 'TRY')"""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return 0.0, ''
    s = str(value).strip()
    if not s:
        return 0.0, ''

    # Para birimini tespit et
    currency = ''
    text_part = re.sub(r'[\d.,\-\s]', '', s).strip()
    if text_part:
        currency = COST_CURRENCY_MAP.get(text_part.upper(), text_part.upper())

    # Sayısal karakterleri, nokta ve virgülü al
    num_str = re.sub(r'[^\d.,\-]', '', s)
    if not num_str:
        return 0.0, currency
    # Türk formatı: 21.500,00 -> 21500.00
    if ',' in num_str and '.' in num_str:
        if num_str.rindex(',') > num_str.rindex('.'):
            num_str = num_str.replace('.', '').replace(',', '.')
        else:
            num_str = num_str.replace(',', '')
    elif ',' in num_str:
        parts = num_str.split(',')
        if len(parts[-1]) == 3 and len(parts) > 1:
            num_str = num_str.replace(',', '')
        else:
            num_str = num_str.replace(',', '.')
    try:
        return float(num_str), currency
    except ValueError:
        return 0.0, currency


class Tooltip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tip_window = None
        self._after_id = None
        widget.bind('<Enter>', self._schedule_show)
        widget.bind('<Leave>', self._hide)

    def _schedule_show(self, event=None):
        self._cancel()
        self._after_id = self.widget.after(400, self._show)

    def _show(self):
        self._after_id = None
        if self.tip_window:
            return
        import tkinter as tk
        x = self.widget.winfo_rootx() + 20
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 5
        self.tip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        tw.attributes('-topmost', True)
        tw.attributes('-disabled', True)
        label = tk.Label(
            tw, text=self.text,
            font=("Segoe UI", 11),
            bg="#34495E", fg="white",
            padx=10, pady=5
        )
        label.pack()

    def _hide(self, event=None):
        self._cancel()
        if self.tip_window:
            self.tip_window.destroy()
            self.tip_window = None

    def _cancel(self):
        if self._after_id:
            self.widget.after_cancel(self._after_id)
            self._after_id = None


class OrderSummaryMerger:
    def __init__(self, root):
        self.root = root
        self.uploaded_files = []
        self.file_item_counts = {}
        self.output_path = None
        self.custom_output_dir = None
        self.is_processing = False
        self._pulsing = False
        self._all_buttons = []

        self._last_browse_dir = self._load_setting('last_browse_dir', '')

        # Openpyxl stil objeleri
        self._thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        self._no_border = Border()
        self._header_fill = PatternFill(start_color='3498DB', end_color='3498DB', fill_type='solid')
        self._header_font = Font(bold=True, size=11, color='FFFFFF')
        self._center_align = Alignment(horizontal='center', vertical='center')
        self._data_align = Alignment(vertical='center', wrap_text=True)
        self._bold_font = Font(bold=True, size=11)
        self._right_align = Alignment(horizontal='right', vertical='center')

        self.setup_ui()
        self._setup_dnd()

    # ── Ayarlar ──────────────────────────────────────────────

    def _load_setting(self, key, default=None):
        try:
            if SETTINGS_FILE.exists():
                with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f).get(key, default)
        except Exception:
            pass
        return default

    def _save_setting(self, key, value):
        try:
            settings = {}
            if SETTINGS_FILE.exists():
                with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
            settings[key] = value
            with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
                json.dump(settings, f)
        except Exception:
            pass

    # ── Drag & Drop ──────────────────────────────────────────

    def _setup_dnd(self):
        if not HAS_DND:
            return
        try:
            self.root.drop_target_register(DND_FILES)
            self.root.dnd_bind('<<Drop>>', self._on_drop)
            self.drop_area.configure(
                text="📂 Dosya Sürükle & Bırak\nveya tıkla"
            )
        except Exception:
            pass

    def _on_drop(self, event):
        files = self.root.tk.splitlist(event.data)
        added = False
        for file_path in files:
            path = Path(file_path)
            if path.suffix.lower() == '.xlsx' and path not in self.uploaded_files:
                self.uploaded_files.append(path)
                added = True
        if added:
            self._scan_and_update()

    # ── UI ───────────────────────────────────────────────────

    def setup_ui(self):
        self.root.title("Sipariş Özeti Birleştirme Aracı")
        self.root.geometry("900x900")
        self.root.minsize(700, 700)

        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)

        main_container = ctk.CTkScrollableFrame(self.root, fg_color="#F8F9FA")
        main_container.grid(row=0, column=0, sticky="nsew")
        main_container.grid_columnconfigure(0, weight=1)

        # HEADER
        header_frame = ctk.CTkFrame(main_container, fg_color="#FFFFFF")
        header_frame.grid(row=0, column=0, sticky="ew", padx=0, pady=0)
        header_frame.grid_columnconfigure(0, weight=1)

        title_container = ctk.CTkFrame(header_frame, fg_color="#FFFFFF")
        title_container.grid(row=0, column=0, sticky="ew", padx=30, pady=20)
        title_container.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(title_container, text="📦", font=("Segoe UI", 32, "bold"), text_color="#2C3E50").grid(row=0, column=0, padx=(0, 15))
        ctk.CTkLabel(title_container, text="Sipariş Özeti Birleştirme Aracı", font=("Segoe UI", 28, "bold"), text_color="#2C3E50").grid(row=0, column=1, sticky="w")
        ctk.CTkLabel(title_container, text="Sipariş özetlerini birleştir (alış fiyatları dahil)", font=("Segoe UI", 12), text_color="#7F8C8D").grid(row=1, column=0, columnspan=2, sticky="w", pady=(5, 0))

        # CONTENT
        content_frame = ctk.CTkFrame(main_container, fg_color="#F8F9FA")
        content_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=20)
        content_frame.grid_columnconfigure(0, weight=1)

        # ── UPLOAD CARD ──
        upload_card = self._create_card(content_frame, "📂 Dosya Seçimi")
        upload_card.grid(row=0, column=0, sticky="ew", pady=(0, 20))

        self.drop_area = ctk.CTkButton(
            upload_card,
            text="👆 Dosya Seçmek İçin Tıkla\n(Ctrl ile birden fazla)",
            command=self.browse_files,
            fg_color="#3498DB",
            hover_color="#2980B9",
            text_color="white",
            font=("Segoe UI", 14, "bold"),
            height=80,
            corner_radius=12
        )
        self.drop_area.pack(fill="both", expand=True, padx=15, pady=15)

        # ── FILE LIST CARD ──
        file_list_card = self._create_card(content_frame, "📋 Seçili Dosyalar")
        file_list_card.grid(row=1, column=0, sticky="nsew", pady=(0, 20))

        tree_frame = ctk.CTkFrame(file_list_card, fg_color="#FFFFFF", corner_radius=8)
        tree_frame.pack(fill="both", expand=True, padx=15, pady=(10, 10))

        style = ttk.Style()
        style.theme_use('clam')
        style.configure('Treeview', rowheight=35, font=("Segoe UI", 11), background="#FFFFFF", foreground="#2C3E50", fieldbackground="#FFFFFF")
        style.configure('Treeview.Heading', font=("Segoe UI", 12, "bold"), background="#ECF0F1", foreground="#2C3E50")
        style.map('Treeview', background=[('selected', '#3498DB')], foreground=[('selected', 'white')])

        self.tree = ttk.Treeview(tree_frame, columns=("name", "items"), show="headings", height=8, selectmode="extended")
        self.tree.heading("name", text="📄 Dosya Adı")
        self.tree.heading("items", text="Durum")
        self.tree.column("name", anchor="w", width=300)
        self.tree.column("items", anchor="center", width=120)
        self.tree.tag_configure('even', background='#F8FBFF')
        self.tree.tag_configure('odd', background='#FFFFFF')

        scrollbar = ttk.Scrollbar(tree_frame, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # BUTTONS
        btn_frame = ctk.CTkFrame(file_list_card, fg_color="#FFFFFF")
        btn_frame.pack(fill="x", padx=15, pady=(0, 15))

        btn_add = ctk.CTkButton(btn_frame, text="➕ Ekle", command=self.browse_files, fg_color="#27AE60", hover_color="#229954", text_color="white", font=("Segoe UI", 11, "bold"), width=100, corner_radius=8)
        btn_add.pack(side="left", padx=(0, 8))
        btn_del = ctk.CTkButton(btn_frame, text="🗑️ Sil", command=self.remove_selected, fg_color="#E74C3C", hover_color="#C0392B", text_color="white", font=("Segoe UI", 11, "bold"), width=100, corner_radius=8)
        btn_del.pack(side="left", padx=4)
        btn_clear = ctk.CTkButton(btn_frame, text="🔄 Temizle", command=self.clear_all, fg_color="#95A5A6", hover_color="#7F8C8D", text_color="white", font=("Segoe UI", 11, "bold"), width=100, corner_radius=8)
        btn_clear.pack(side="left", padx=4)

        ctk.CTkFrame(btn_frame, fg_color="#ECF0F1", width=2).pack(side="left", fill="y", padx=10, pady=2)

        btn_up = ctk.CTkButton(btn_frame, text="🔼", command=self.move_up, fg_color="#8E44AD", hover_color="#7D3C98", text_color="white", font=("Segoe UI", 11, "bold"), width=50, corner_radius=8)
        btn_up.pack(side="left", padx=4)
        btn_down = ctk.CTkButton(btn_frame, text="🔽", command=self.move_down, fg_color="#8E44AD", hover_color="#7D3C98", text_color="white", font=("Segoe UI", 11, "bold"), width=50, corner_radius=8)
        btn_down.pack(side="left", padx=4)

        Tooltip(btn_add, "Yeni dosya ekle")
        Tooltip(btn_del, "Seçili dosyaları sil")
        Tooltip(btn_clear, "Tüm listeyi temizle")
        Tooltip(btn_up, "Seçili dosyayı yukarı taşı")
        Tooltip(btn_down, "Seçili dosyayı aşağı taşı")

        # ── DISCOUNT CARD ──
        discount_card = self._create_card(content_frame, "💰 Firma İndirim Oranı")
        discount_card.grid(row=2, column=0, sticky="ew", pady=(0, 20))

        discount_inner = ctk.CTkFrame(discount_card, fg_color="#FFFFFF")
        discount_inner.pack(fill="x", padx=15, pady=(10, 15))

        ctk.CTkLabel(
            discount_inner,
            text="İndirim Oranı (%):",
            font=("Segoe UI", 12),
            text_color="#2C3E50"
        ).pack(side="left", padx=(0, 10))

        saved_discount = self._load_setting('discount_pct', '0')
        self.discount_var = ctk.StringVar(value=str(saved_discount))
        self.discount_entry = ctk.CTkEntry(
            discount_inner,
            textvariable=self.discount_var,
            width=80,
            font=("Segoe UI", 13, "bold"),
            justify="center",
            placeholder_text="0"
        )
        self.discount_entry.pack(side="left", padx=(0, 10))

        ctk.CTkLabel(
            discount_inner,
            text="%",
            font=("Segoe UI", 14, "bold"),
            text_color="#E74C3C"
        ).pack(side="left")

        ctk.CTkLabel(
            discount_inner,
            text="  (Sipariş özetlerinde belirtilmeyen firma indirimi)",
            font=("Segoe UI", 10),
            text_color="#95A5A6"
        ).pack(side="left", padx=(10, 0))

        # ── EXCHANGE RATE CARD ──
        fx_card = self._create_card(content_frame, "💱 Döviz Kurları (Alış = Satış Kuru)")
        fx_card.grid(row=3, column=0, sticky="ew", pady=(0, 20))

        fx_inner = ctk.CTkFrame(fx_card, fg_color="#FFFFFF")
        fx_inner.pack(fill="x", padx=15, pady=(10, 15))

        # Kur güncelle butonu + durum
        fx_top = ctk.CTkFrame(fx_inner, fg_color="#FFFFFF")
        fx_top.pack(fill="x", pady=(0, 10))

        self.fx_fetch_btn = ctk.CTkButton(
            fx_top, text="🔄 Güncel Kurları Çek",
            command=self._fetch_rates_async,
            fg_color="#E67E22", hover_color="#D35400",
            text_color="white", font=("Segoe UI", 12, "bold"),
            width=200, height=35, corner_radius=8
        )
        self.fx_fetch_btn.pack(side="left", padx=(0, 10))

        self.fx_status_label = ctk.CTkLabel(
            fx_top, text="",
            font=("Segoe UI", 10), text_color="#95A5A6"
        )
        self.fx_status_label.pack(side="left")

        # Kur input'ları
        self.fx_entries = {}

        for code, label_text, default in [('eur', '1 EUR =', '38.50'), ('usd', '1 USD =', '36.20')]:
            frame = ctk.CTkFrame(fx_inner, fg_color="#FFFFFF")
            frame.pack(fill="x", pady=(0, 6))
            ctk.CTkLabel(frame, text=label_text, font=("Segoe UI", 12, "bold"), text_color="#2C3E50", width=80).pack(side="left", padx=(0, 8))
            saved = self._load_setting(f'{code}_tl_rate', default)
            var = ctk.StringVar(value=str(saved))
            entry = ctk.CTkEntry(frame, textvariable=var, width=100, font=("Segoe UI", 13, "bold"), justify="center", placeholder_text=default)
            entry.pack(side="left", padx=(0, 5))
            ctk.CTkLabel(frame, text="TL", font=("Segoe UI", 12, "bold"), text_color="#E67E22").pack(side="left")
            self.fx_entries[code] = var

        self.eur_tl_var = self.fx_entries['eur']
        self.usd_tl_var = self.fx_entries['usd']

        ctk.CTkLabel(
            fx_inner,
            text="Kurlar online çekilir, manuel de düzenlenebilir",
            font=("Segoe UI", 10),
            text_color="#95A5A6"
        ).pack(anchor="w", pady=(5, 0))

        # Başlangıçta son güncelleme zamanını göster
        last_update = self._load_setting('fx_last_update', '')
        if last_update:
            self.fx_status_label.configure(text=f"Son: {last_update}", text_color="#27AE60")

        # ── OUTPUT PATH CARD ──
        output_card = self._create_card(content_frame, "📁 Çıktı Konumu")
        output_card.grid(row=4, column=0, sticky="ew", pady=(0, 20))

        output_inner = ctk.CTkFrame(output_card, fg_color="#FFFFFF")
        output_inner.pack(fill="x", padx=15, pady=(10, 15))
        output_inner.grid_columnconfigure(0, weight=1)

        self.output_dir_label = ctk.CTkLabel(
            output_inner,
            text="İlk dosyanın klasörü (varsayılan)",
            font=("Segoe UI", 11),
            text_color="#7F8C8D",
            anchor="w"
        )
        self.output_dir_label.grid(row=0, column=0, sticky="ew", padx=(0, 10))

        btn_out = ctk.CTkButton(output_inner, text="📂 Seç", command=self.choose_output_dir, fg_color="#34495E", hover_color="#2C3E50", text_color="white", font=("Segoe UI", 11, "bold"), width=80, corner_radius=8)
        btn_out.grid(row=0, column=1)
        btn_reset = ctk.CTkButton(output_inner, text="↺", command=self.reset_output_dir, fg_color="#95A5A6", hover_color="#7F8C8D", text_color="white", font=("Segoe UI", 11, "bold"), width=40, corner_radius=8)
        btn_reset.grid(row=0, column=2, padx=(5, 0))
        Tooltip(btn_out, "Çıktı klasörü seç")
        Tooltip(btn_reset, "Varsayılana sıfırla")

        # ── STATUS CARD ──
        status_card = self._create_card(content_frame, "⚙️ Durum")
        status_card.grid(row=5, column=0, sticky="ew", pady=(0, 20))

        self.status_label = ctk.CTkLabel(status_card, text="✅ Hazır", font=("Segoe UI", 12), text_color="#27AE60")
        self.status_label.pack(anchor="w", padx=15, pady=(10, 5))

        self.progress = ctk.CTkProgressBar(status_card, fg_color="#ECF0F1", progress_color="#3498DB", height=6, corner_radius=3)
        self.progress.pack(fill="x", padx=15, pady=(5, 15))
        self.progress.set(0)

        # ── OPTIONS ──
        options_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
        options_frame.grid(row=6, column=0, sticky="ew", pady=(0, 10))

        self.auto_open_var = ctk.BooleanVar(value=self._load_setting('auto_open', False))
        ctk.CTkCheckBox(
            options_frame,
            text="Bitince dosyayı otomatik aç",
            variable=self.auto_open_var,
            font=("Segoe UI", 12),
            text_color="#2C3E50",
            command=lambda: self._save_setting('auto_open', self.auto_open_var.get())
        ).pack(anchor="w")

        self.show_header_info_var = ctk.BooleanVar(value=self._load_setting('show_header_info', True))
        ctk.CTkCheckBox(
            options_frame,
            text="Sipariş bilgilerini göster (Tarih, RFQ, QTN)",
            variable=self.show_header_info_var,
            font=("Segoe UI", 12),
            text_color="#2C3E50",
            command=lambda: self._save_setting('show_header_info', self.show_header_info_var.get())
        ).pack(anchor="w", pady=(5, 0))

        # ── ACTION BUTTONS ──
        action_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
        action_frame.grid(row=7, column=0, sticky="ew", pady=(10, 0))
        action_frame.grid_columnconfigure(0, weight=1)
        action_frame.grid_columnconfigure(1, weight=1)

        self.merge_btn = ctk.CTkButton(
            action_frame,
            text="🚀 Dosyaları Birleştir",
            command=self.merge_files,
            fg_color="#2980B9",
            hover_color="#1F618D",
            text_color="white",
            font=("Segoe UI", 14, "bold"),
            height=50,
            corner_radius=10,
            state="disabled"
        )
        self.merge_btn.grid(row=0, column=0, sticky="ew", padx=(0, 10))

        self.open_btn = ctk.CTkButton(
            action_frame,
            text="📄 Sonucu Aç",
            command=self.open_file,
            fg_color="#16A085",
            hover_color="#117A65",
            text_color="white",
            font=("Segoe UI", 14, "bold"),
            height=50,
            corner_radius=10,
            state="disabled"
        )
        self.open_btn.grid(row=0, column=1, sticky="ew")

        Tooltip(self.drop_area, "Sipariş özeti Excel dosyalarını seçmek için tıkla")
        Tooltip(self.merge_btn, "Seçili dosyaları tek bir Excel'de birleştir")
        Tooltip(self.open_btn, "Oluşturulan birleştirilmiş dosyayı aç")

        self._all_buttons = [
            self.drop_area, btn_add, btn_del, btn_clear,
            btn_up, btn_down, btn_out, btn_reset,
            self.merge_btn, self.open_btn
        ]

    def _create_card(self, parent, title):
        card = ctk.CTkFrame(parent, fg_color="#FFFFFF", corner_radius=12, border_width=1, border_color="#E8F4F8")
        ctk.CTkLabel(card, text=title, font=("Segoe UI", 13, "bold"), text_color="#2C3E50").pack(anchor="w", padx=15, pady=(15, 0))
        ctk.CTkFrame(card, fg_color="#ECF0F1", height=1).pack(fill="x", padx=15, pady=(10, 0))
        return card

    # ── Dosya İşlemleri ──────────────────────────────────────

    def browse_files(self):
        initial_dir = self._last_browse_dir if self._last_browse_dir else None
        files = filedialog.askopenfilenames(
            title="Sipariş Özeti Dosyalarını Seçin",
            initialdir=initial_dir,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not files:
            return
        added = False
        for file_path in files:
            path = Path(file_path)
            if path.suffix.lower() == '.xlsx' and path not in self.uploaded_files:
                self.uploaded_files.append(path)
                added = True
        self._last_browse_dir = str(Path(files[0]).parent)
        self._save_setting('last_browse_dir', self._last_browse_dir)
        if added:
            self._scan_and_update()

    def _scan_and_update(self):
        self.update_file_list()
        threading.Thread(target=self._scan_worker, daemon=True).start()

    def _scan_worker(self):
        for f in list(self.uploaded_files):
            if f not in self.file_item_counts:
                data = self._extract_order_data(f)
                self.file_item_counts[f] = len(data['data_rows']) if data else -1
        self.root.after(0, self.update_file_list)

    def update_file_list(self):
        self.tree.delete(*self.tree.get_children())
        for i, f in enumerate(self.uploaded_files):
            count = self.file_item_counts.get(f)
            if count is None:
                status = "⏳ Taranıyor..."
            elif count < 0:
                status = "⚠️ Okunamadı"
            else:
                status = f"📊 {count} item"
            tag = 'even' if i % 2 == 0 else 'odd'
            self.tree.insert("", "end", values=(f.name, status), tags=(tag,))

        file_count = len(self.uploaded_files)
        if file_count > 0:
            total = sum(c for c in self.file_item_counts.values() if c and c > 0)
            self.status_label.configure(
                text=f"✅ {file_count} dosya seçildi ({total} item)" if total else f"✅ {file_count} dosya seçildi",
                text_color="#27AE60"
            )
            self.merge_btn.configure(state="normal")
        else:
            self.status_label.configure(text="⏳ Dosya seçin", text_color="#7F8C8D")
            self.merge_btn.configure(state="disabled")

    def remove_selected(self):
        selected = self.tree.selection()
        if not selected:
            return
        indices = sorted([self.tree.index(item) for item in selected], reverse=True)
        for idx in indices:
            if 0 <= idx < len(self.uploaded_files):
                removed = self.uploaded_files.pop(idx)
                self.file_item_counts.pop(removed, None)
        self.update_file_list()

    def clear_all(self):
        self.uploaded_files.clear()
        self.file_item_counts.clear()
        self.update_file_list()
        self.open_btn.configure(state="disabled")

    def move_up(self):
        selected = self.tree.selection()
        if not selected or len(selected) != 1:
            return
        idx = self.tree.index(selected[0])
        if idx > 0:
            self.uploaded_files[idx], self.uploaded_files[idx - 1] = self.uploaded_files[idx - 1], self.uploaded_files[idx]
            self.update_file_list()
            self.tree.selection_set(self.tree.get_children()[idx - 1])

    def move_down(self):
        selected = self.tree.selection()
        if not selected or len(selected) != 1:
            return
        idx = self.tree.index(selected[0])
        if idx < len(self.uploaded_files) - 1:
            self.uploaded_files[idx], self.uploaded_files[idx + 1] = self.uploaded_files[idx + 1], self.uploaded_files[idx]
            self.update_file_list()
            self.tree.selection_set(self.tree.get_children()[idx + 1])

    # ── Çıktı Konumu ─────────────────────────────────────────

    def choose_output_dir(self):
        dir_path = filedialog.askdirectory(title="Çıktı Klasörünü Seçin")
        if dir_path:
            self.custom_output_dir = Path(dir_path)
            self.output_dir_label.configure(text=str(self.custom_output_dir), text_color="#2C3E50")

    def reset_output_dir(self):
        self.custom_output_dir = None
        self.output_dir_label.configure(text="İlk dosyanın klasörü (varsayılan)", text_color="#7F8C8D")

    # ── Birleştirme ──────────────────────────────────────────

    def _get_discount_pct(self):
        try:
            val = float(self.discount_var.get().replace(',', '.').strip())
            self._save_setting('discount_pct', val)
            return val
        except (ValueError, TypeError):
            return 0.0

    def _get_fx_rates(self):
        """Döviz kurlarını al ve kaydet. TL cinsinden kurları döndürür."""
        rates = {}
        for key, var, default in [
            ('eur_tl_rate', self.eur_tl_var, 38.50),
            ('usd_tl_rate', self.usd_tl_var, 36.20),
        ]:
            try:
                val = float(var.get().replace(',', '.').strip())
            except (ValueError, TypeError):
                val = default
            self._save_setting(key, val)
            rates[key] = val
        return {
            'TRY': 1.0,
            'EUR': rates['eur_tl_rate'],
            'USD': rates['usd_tl_rate'],
        }

    def _fetch_rates_async(self):
        """Kurları online çek (arka planda)"""
        self.fx_fetch_btn.configure(state="disabled", text="⏳ Çekiliyor...")
        self.fx_status_label.configure(text="Bağlanıyor...", text_color="#F39C12")
        threading.Thread(target=self._fetch_rates_worker, daemon=True).start()

    def _fetch_rates_worker(self):
        """Frankfurter API'den (ECB verileri) güncel kurları çek"""
        try:
            url = "https://api.frankfurter.app/latest?from=TRY&to=EUR,USD"
            req = urllib.request.Request(url, headers={'User-Agent': 'OrderMerger/1.0'})
            with urllib.request.urlopen(req, timeout=10) as resp:
                data = json.loads(resp.read().decode('utf-8'))

            # API: 1 TRY = X EUR, 1 TRY = Y USD  ->  1 EUR = 1/X TL, 1 USD = 1/Y TL
            rates = data.get('rates', {})
            eur_rate = rates.get('EUR')
            usd_rate = rates.get('USD')

            if eur_rate and usd_rate:
                eur_tl = round(1.0 / eur_rate, 4)
                usd_tl = round(1.0 / usd_rate, 4)

                now = datetime.now().strftime('%d.%m.%Y %H:%M')

                def _update():
                    self.eur_tl_var.set(str(eur_tl))
                    self.usd_tl_var.set(str(usd_tl))
                    self._save_setting('eur_tl_rate', eur_tl)
                    self._save_setting('usd_tl_rate', usd_tl)
                    self._save_setting('fx_last_update', now)
                    self.fx_status_label.configure(
                        text=f"Guncellendi: {now} (ECB)",
                        text_color="#27AE60"
                    )
                    self.fx_fetch_btn.configure(state="normal", text="🔄 Güncel Kurları Çek")

                self.root.after(0, _update)
            else:
                raise ValueError("Kur verisi alınamadı")

        except Exception as e:
            err = str(e)

            def _error():
                self.fx_status_label.configure(
                    text=f"Hata: {err[:50]}",
                    text_color="#E74C3C"
                )
                self.fx_fetch_btn.configure(state="normal", text="🔄 Güncel Kurları Çek")

            self.root.after(0, _error)

    def _convert_cost(self, amount, cost_currency, sale_currency, fx_rates):
        """Cost'u satış para birimine çevir. Tüm kurlar TL cinsindendir."""
        if not cost_currency or cost_currency == sale_currency:
            return amount
        cost_in_tl = amount * fx_rates.get(cost_currency, 1.0)
        sale_rate = fx_rates.get(sale_currency, 1.0)
        if sale_rate == 0:
            return amount
        return cost_in_tl / sale_rate

    def merge_files(self):
        if self.is_processing:
            return
        self.is_processing = True
        self._lock_ui()
        threading.Thread(target=self._merge_worker, daemon=True).start()

    def _check_write_permission(self, dir_path):
        try:
            test_file = dir_path / '.write_test_tmp'
            test_file.touch()
            test_file.unlink()
            return True
        except Exception:
            return False

    def _is_file_locked(self, file_path):
        if not file_path.exists():
            return False
        try:
            with open(file_path, 'r+b'):
                return False
        except (IOError, PermissionError):
            return True

    def _merge_worker(self):
        try:
            self._update_progress(0)
            self._update_status("⏳ Hazırlanıyor...", "#F39C12")

            output_dir = self.custom_output_dir or self.uploaded_files[0].parent
            if not self._check_write_permission(output_dir):
                self._update_status("❌ Hata!", "#E74C3C")
                err_dir = str(output_dir)
                self.root.after(0, lambda: messagebox.showerror(
                    "Hata",
                    f"Çıktı klasörüne yazılamıyor!\n{err_dir}"
                ))
                return

            self._update_progress(0.15)

            self._update_progress(0.3)
            self._update_status("📊 Dosyalar birleştiriliyor... (%30)", "#F39C12")
            self.root.after(0, self._start_pulse)

            discount_pct = self._get_discount_pct()
            fx_rates = self._get_fx_rates()
            total_items, vessel_names = self._create_merged_file(discount_pct, fx_rates)

            # Dosya adı: GemiIsmi_tarih.xlsx
            date_str = datetime.now().strftime('%d-%m-%Y')
            if vessel_names:
                vessel_str = '_'.join(vessel_names)
                vessel_str = re.sub(r'[<>:"/\\|?*]', '', vessel_str).strip()
                filename = f'{vessel_str}_{date_str}.xlsx'
            else:
                filename = f'MERGED_ORDER_SUMMARY_{date_str}.xlsx'
            self.output_path = output_dir / filename
            self._pending_wb.save(self.output_path)
            del self._pending_wb

            self.root.after(0, lambda: self._stop_pulse(0.9))
            self._update_progress(1.0)

            file_count = len(self.uploaded_files)
            disc_text = f", İndirim: %{discount_pct}" if discount_pct > 0 else ""
            self._update_status(f"✅ Tamamlandı! ({file_count} sipariş, {total_items} item{disc_text})", "#27AE60")

            if self.auto_open_var.get():
                self.root.after(0, self.open_file)
            else:
                out_name = self.output_path.name
                out_parent = str(self.output_path.parent)
                self.root.after(0, lambda: messagebox.showinfo(
                    "✅ Başarılı",
                    f"Sipariş Özeti oluşturuldu!\n\n📁 {out_name}\n📍 {out_parent}\n\n📊 {file_count} sipariş\n🔢 {total_items} item{disc_text}"
                ))

            self.root.after(300, self._show_verification_warning)

        except Exception as e:
            self.root.after(0, lambda: self._stop_pulse(0))
            self._update_status("❌ Hata!", "#E74C3C")
            error_msg = str(e)
            self.root.after(0, lambda: messagebox.showerror("Hata", f"Birleştirme hatası:\n{error_msg}"))
        finally:
            self.is_processing = False
            self.root.after(0, self._unlock_ui)

    def _show_verification_warning(self):
        dlg = ctk.CTkToplevel(self.root)
        dlg.title("⚠️ Önemli Uyarı")
        dlg.geometry("520x280")
        dlg.resizable(False, False)
        dlg.attributes('-topmost', True)
        dlg.grab_set()

        dlg.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() - 520) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - 280) // 2
        dlg.geometry(f"+{x}+{y}")

        ctk.CTkLabel(dlg, text="⚠️", font=("Segoe UI", 48)).pack(pady=(20, 5))
        ctk.CTkLabel(
            dlg,
            text="TOPLAM TUTARLARI MUTLAKA\nELLE KONTROL EDİNİZ!",
            font=("Segoe UI", 18, "bold"),
            text_color="#C0392B"
        ).pack(pady=(0, 5))
        ctk.CTkLabel(
            dlg,
            text="Birleştirilmiş dosyadaki tüm fiyat ve toplam değerlerini\ngöndermeden önce manuel olarak doğrulamanız gerekmektedir.",
            font=("Segoe UI", 11),
            text_color="#7F8C8D"
        ).pack(pady=(0, 15))

        btn = ctk.CTkButton(
            dlg, text="Anladım (3)",
            fg_color="#95A5A6", hover_color="#95A5A6",
            text_color="white", font=("Segoe UI", 13, "bold"),
            width=200, height=40, corner_radius=8, state="disabled"
        )
        btn.pack(pady=(0, 20))

        def countdown(sec):
            if sec > 0:
                btn.configure(text=f"Anladım ({sec})")
                dlg.after(1000, countdown, sec - 1)
            else:
                btn.configure(
                    text="Anladım ✓", state="normal",
                    fg_color="#27AE60", hover_color="#229954",
                    command=dlg.destroy
                )

        countdown(3)

    def _update_status(self, text, color):
        self.root.after(0, lambda: self.status_label.configure(text=text, text_color=color))

    def _update_progress(self, value):
        self.root.after(0, lambda: self.progress.set(value))

    def _start_pulse(self):
        self._pulsing = True
        self._pulse_val = 0.0
        self._pulse_dir = 0.02
        self._do_pulse()

    def _do_pulse(self):
        if not self._pulsing:
            return
        self._pulse_val += self._pulse_dir
        if self._pulse_val >= 1.0 or self._pulse_val <= 0.0:
            self._pulse_dir *= -1
        self.progress.set(self._pulse_val)
        self.root.after(30, self._do_pulse)

    def _stop_pulse(self, final_value=1.0):
        self._pulsing = False
        self.progress.set(final_value)

    def _lock_ui(self):
        for btn in self._all_buttons:
            btn.configure(state="disabled")

    def _unlock_ui(self):
        for btn in self._all_buttons:
            btn.configure(state="normal")
        if not self.uploaded_files:
            self.merge_btn.configure(state="disabled")
        if not (self.output_path and self.output_path.exists()):
            self.open_btn.configure(state="disabled")

    # ── Excel İşlemleri ──────────────────────────────────────

    def _create_merged_file(self, discount_pct, fx_rates):
        wb = Workbook()
        ws = wb.active
        ws.title = "Merged Order Summary"

        headers = ['NO', 'DESCRIPTION', 'CODE', 'QTTY', 'UNIT', 'U.PRICE', 'T.PRICE', 'REMARKS', 'STOCK LOC.', 'U.COST', 'T.COST']
        total_cols = 11

        # Önce tüm dosyaları oku (gemi isimlerini toplamak için)
        all_orders = []
        vessel_names = []
        for file_path in self.uploaded_files:
            order_data = self._extract_order_data(file_path)
            if not order_data:
                continue
            all_orders.append(order_data)
            v = order_data['header_info'].get('vessel', '')
            if v and v not in vessel_names:
                vessel_names.append(v)

        # ── EXCEL HEADER BANNER ──
        current_row = 1

        # Row 1: Ana başlık banner (koyu lacivert)
        banner_fill = PatternFill(start_color='1B2631', end_color='1B2631', fill_type='solid')
        banner_font = Font(bold=True, size=18, color='FFFFFF')
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title_cell = ws.cell(1, 1)
        title_cell.value = 'MERGED ORDER SUMMARY'
        title_cell.font = banner_font
        title_cell.fill = banner_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        for c in range(2, total_cols + 1):
            ws.cell(1, c).fill = banner_fill
        ws.row_dimensions[1].height = 40

        # Row 2: Gemi ismi satırı (koyu gri-mavi)
        vessel_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        vessel_font = Font(bold=True, size=13, color='F39C12')
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
        vessel_cell = ws.cell(2, 1)
        vessel_text = ' / '.join(vessel_names) if vessel_names else 'N/A'
        vessel_cell.value = f'VESSEL: {vessel_text}'
        vessel_cell.font = vessel_font
        vessel_cell.fill = vessel_fill
        vessel_cell.alignment = Alignment(horizontal='center', vertical='center')
        for c in range(2, total_cols + 1):
            ws.cell(2, c).fill = vessel_fill
        ws.row_dimensions[2].height = 30

        # Row 3: Alt bilgi satırı (mavi accent)
        accent_fill = PatternFill(start_color='2980B9', end_color='2980B9', fill_type='solid')
        accent_font = Font(bold=True, size=10, color='FFFFFF')
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_cols)
        info_cell = ws.cell(3, 1)
        now_str = datetime.now().strftime('%d.%m.%Y %H:%M')
        file_count = len(all_orders)
        disc_info = f'  |  Discount: %{discount_pct}' if discount_pct > 0 else ''
        info_cell.value = f'Generated: {now_str}  |  Files: {file_count}{disc_info}'
        info_cell.font = accent_font
        info_cell.fill = accent_fill
        info_cell.alignment = Alignment(horizontal='center', vertical='center')
        for c in range(2, total_cols + 1):
            ws.cell(3, c).fill = accent_fill
        ws.row_dimensions[3].height = 24

        # Row 4: İnce ayırıcı çizgi (altın sarısı)
        gold_fill = PatternFill(start_color='F39C12', end_color='F39C12', fill_type='solid')
        for c in range(1, total_cols + 1):
            ws.cell(4, c).fill = gold_fill
        ws.row_dimensions[4].height = 4

        current_row = 6
        total_items = 0

        all_sale_total_rows = []
        all_cost_total_rows = []
        last_currency_symbol = ''

        for order_data in all_orders:

            info = order_data['header_info']
            sale_currency = info.get('currency', '').upper()
            currency_symbol = CURRENCY_SYMBOLS.get(sale_currency, sale_currency) if sale_currency else ''
            if currency_symbol:
                last_currency_symbol = currency_symbol

            # Sipariş bilgi satırı
            info_text = f"Order: {order_data['file_name']}"
            header_cells = order_data.get('header_cells', [])
            show_cells = self.show_header_info_var.get()
            has_cells = show_cells and any(l or v for l, v in header_cells)

            if has_cells:
                info_label_font = Font(bold=True, size=9)
                info_value_font = Font(size=9)
                ws.cell(current_row, 2).value = info_text
                ws.cell(current_row, 2).font = Font(italic=True, size=9, color='808080')
                for i, (label, value) in enumerate(header_cells):
                    r = current_row + i
                    if label or value:
                        cell_h = ws.cell(r, 8)
                        clean_label = label.rstrip(' :')
                        cell_h.value = f"{clean_label} : " if clean_label else ''
                        cell_h.font = info_label_font
                        cell_h.alignment = self._right_align
                        cell_h.border = self._thin_border
                        cell_i = ws.cell(r, 9)
                        cell_i.value = value
                        cell_i.font = info_value_font
                        cell_i.border = self._thin_border
                current_row += len(header_cells)
            else:
                ws.cell(current_row, 2).value = info_text
                ws.cell(current_row, 2).font = Font(italic=True, size=9, color='808080')
                current_row += 1

            # Header satırı
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(current_row, col_idx).value = header
            self._apply_header_style(ws, current_row)
            current_row += 1

            # Data satırları
            item_count = 0
            data_start_row = current_row
            price_format = f'"{currency_symbol}"#,##0.00' if currency_symbol else '#,##0.00'

            for data_row in order_data['data_rows']:
                item_count += 1
                # Orijinal veri sütunları: 0=NO, 1=DESC, 2=CODE, 3=QTTY, 4=UNIT, 5=U.PRICE, 6=T.PRICE, 7=REMARKS, 8=STOCK LOC, 9=COST
                qtty = data_row[3] if len(data_row) > 3 and pd.notna(data_row[3]) else 0
                try:
                    qtty_num = float(qtty)
                except (ValueError, TypeError):
                    qtty_num = 0

                # Cost parse + kur dönüşümü
                raw_cost = data_row[9] if len(data_row) > 9 else None
                unit_cost_raw, cost_currency = _parse_cost(raw_cost)
                unit_cost_converted = self._convert_cost(unit_cost_raw, cost_currency, sale_currency, fx_rates) if unit_cost_raw > 0 else 0.0

                for col_idx in range(1, total_cols + 1):
                    cell = ws.cell(current_row, col_idx)

                    if col_idx == 1:
                        cell.value = item_count
                    elif col_idx == 7:
                        # T.PRICE = QTTY * U.PRICE
                        cell.value = f"=D{current_row}*F{current_row}"
                        cell.number_format = price_format
                    elif col_idx == 10:
                        # U.COST (satış para birimine çevrilmiş birim maliyet)
                        cell.value = round(unit_cost_converted, 2) if unit_cost_converted > 0 else None
                        cell.number_format = price_format
                    elif col_idx == 11:
                        # T.COST = QTTY * U.COST
                        cell.value = f"=D{current_row}*J{current_row}"
                        cell.number_format = price_format
                    elif col_idx == 6:
                        value = data_row[col_idx - 1] if col_idx - 1 < len(data_row) else None
                        cell.value = value
                        if value is not None:
                            cell.number_format = price_format
                    else:
                        # Sütun 1-9 arası (COST hariç) orijinal veriyi yaz
                        if col_idx <= 9:
                            value = data_row[col_idx - 1] if col_idx - 1 < len(data_row) else None
                            cell.value = value

                self._apply_data_row_style(ws, current_row)
                current_row += 1

            total_items += item_count
            current_row += 1

            # TOTAL satırı (satış)
            self._apply_total_style(ws, current_row, 'TOTAL:', col_label=6, col_value=7)
            ws.cell(current_row, 7).value = f"=SUM(G{data_start_row}:G{data_start_row + item_count - 1})"
            ws.cell(current_row, 7).number_format = price_format
            all_sale_total_rows.append(current_row)

            # COST TOTAL (alış - satış para biriminde)
            self._apply_total_style(ws, current_row, 'COST TOTAL:', col_label=10, col_value=11)
            ws.cell(current_row, 11).value = f"=SUM(K{data_start_row}:K{data_start_row + item_count - 1})"
            ws.cell(current_row, 11).number_format = price_format
            all_cost_total_rows.append(current_row)
            current_row += 3

        # ── GRAND SUMMARY ──
        if all_sale_total_rows:
            summary_format = f'"{last_currency_symbol}"#,##0.00' if last_currency_symbol else '#,##0.00'

            # Ayırıcı çizgi
            separator_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
            for col in range(1, total_cols + 1):
                cell = ws.cell(current_row, col)
                cell.fill = separator_fill
                cell.border = self._thin_border
            current_row += 1

            # Başlık
            banner_fill = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')
            banner_font = Font(bold=True, size=13, color='FFFFFF')
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
            title_cell = ws.cell(current_row, 1)
            disc_label = f"  |  İNDİRİM: %{discount_pct}" if discount_pct > 0 else ""
            title_cell.value = f'GRAND SUMMARY  —  {len(all_sale_total_rows)} ORDERS{disc_label}'
            title_cell.font = banner_font
            title_cell.fill = banner_fill
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            title_cell.border = self._thin_border
            for col in range(2, total_cols + 1):
                ws.cell(current_row, col).fill = banner_fill
                ws.cell(current_row, col).border = self._thin_border
            current_row += 2

            summary_label_font = Font(bold=True, size=12, color='2C3E50')
            summary_value_font = Font(bold=True, size=12, color='1A5276')
            summary_border = Border(
                left=Side(style='medium'), right=Side(style='medium'),
                top=Side(style='medium'), bottom=Side(style='medium')
            )
            label_fill = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')
            value_fill = PatternFill(start_color='D4E6F1', end_color='D4E6F1', fill_type='solid')

            def _write_summary_row(label, formula, fmt, is_grand=False):
                nonlocal current_row
                if is_grand:
                    fill_l = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')
                    fill_v = fill_l
                    font_l = Font(bold=True, size=14, color='FFFFFF')
                    font_v = font_l
                else:
                    fill_l = label_fill
                    fill_v = value_fill
                    font_l = summary_label_font
                    font_v = summary_value_font

                ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
                lbl = ws.cell(current_row, 4)
                lbl.value = label
                lbl.font = font_l
                lbl.alignment = Alignment(horizontal='right', vertical='center')
                lbl.fill = fill_l
                lbl.border = summary_border
                for c in range(5, 7):
                    ws.cell(current_row, c).fill = fill_l
                    ws.cell(current_row, c).border = summary_border

                ws.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=8)
                val = ws.cell(current_row, 7)
                val.value = formula
                val.font = font_v
                val.number_format = fmt
                val.alignment = Alignment(horizontal='center', vertical='center')
                val.fill = fill_v
                val.border = summary_border
                ws.cell(current_row, 8).fill = fill_v
                ws.cell(current_row, 8).border = summary_border
                current_row += 1

            # TOPLAM SATIŞ
            sale_refs = '+'.join([f'G{r}' for r in all_sale_total_rows])
            _write_summary_row('TOPLAM SATIŞ :', f'={sale_refs}', summary_format)
            sale_total_row = current_row - 1

            # TOPLAM ALIŞ (satış para biriminde)
            cost_refs = '+'.join([f'K{r}' for r in all_cost_total_rows])
            _write_summary_row('TOPLAM ALIŞ :', f'={cost_refs}', summary_format)
            cost_total_row = current_row - 1

            # İNDİRİM (eğer varsa)
            if discount_pct > 0:
                _write_summary_row(
                    f'İNDİRİM ({discount_pct}%) :',
                    f'=G{sale_total_row}*{discount_pct/100}',
                    summary_format
                )
                disc_row = current_row - 1

                # FİNAL SATIŞ TUTARI
                _write_summary_row(
                    'FİNAL SATIŞ TUTARI :',
                    f'=G{sale_total_row}-G{disc_row}',
                    summary_format,
                    is_grand=True
                )
                final_sale_row = current_row - 1
            else:
                final_sale_row = sale_total_row

            # KÂR / ZARAR
            current_row += 1
            profit_fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
            profit_font = Font(bold=True, size=14, color='FFFFFF')
            ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
            lbl = ws.cell(current_row, 4)
            lbl.value = 'KÂR / ZARAR :'
            lbl.font = profit_font
            lbl.alignment = Alignment(horizontal='right', vertical='center')
            lbl.fill = profit_fill
            lbl.border = summary_border
            for c in range(5, 7):
                ws.cell(current_row, c).fill = profit_fill
                ws.cell(current_row, c).border = summary_border

            ws.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=8)
            val = ws.cell(current_row, 7)
            val.value = f'=G{final_sale_row}-G{cost_total_row}'
            val.font = profit_font
            val.number_format = summary_format
            val.alignment = Alignment(horizontal='center', vertical='center')
            val.fill = profit_fill
            val.border = summary_border
            ws.cell(current_row, 8).fill = profit_fill
            ws.cell(current_row, 8).border = summary_border
            current_row += 2

        # Sütun genişlikleri
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 55
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 30
        ws.column_dimensions['I'].width = 18
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 14

        last_row = current_row - 1
        ws.print_area = f'A1:K{last_row}'
        ws.sheet_view.showGridLines = False

        self._pending_wb = wb
        return total_items, vessel_names

    def _extract_order_data(self, file_path):
        try:
            df = pd.read_excel(file_path, header=None)

            header_info = {}
            if len(df.columns) < 2:
                return None

            # Gemi ismi: 15B hücresi (0-indexed: row 14, col 1)
            vessel_name = ''
            if len(df) > 14 and len(df.columns) > 1 and pd.notna(df.iloc[14, 1]):
                vessel_name = str(df.iloc[14, 1]).strip()
            header_info['vessel'] = vessel_name

            # Header bilgileri: satır 18-20 (0-indexed: 17-19), col 7-8 (0-indexed: 7-8)
            header_cells = []
            for row_idx in range(min(25, len(df))):
                for col_idx in range(min(10, len(df.columns))):
                    cell_val = str(df.iloc[row_idx, col_idx]).strip() if pd.notna(df.iloc[row_idx, col_idx]) else ''

                    if 'RFQ REF' in cell_val.upper():
                        # Değer yanındaki sütunda
                        next_col = col_idx + 1
                        if next_col < len(df.columns) and pd.notna(df.iloc[row_idx, next_col]):
                            header_info['rfq_ref'] = str(df.iloc[row_idx, next_col]).strip()
                    elif 'QTN REF' in cell_val.upper():
                        next_col = col_idx + 1
                        if next_col < len(df.columns) and pd.notna(df.iloc[row_idx, next_col]):
                            header_info['qtn_ref'] = str(df.iloc[row_idx, next_col]).strip()
                    elif 'DATE' in cell_val.upper() and ':' in cell_val:
                        next_col = col_idx + 1
                        if next_col < len(df.columns) and pd.notna(df.iloc[row_idx, next_col]):
                            header_info['date'] = str(df.iloc[row_idx, next_col]).strip()

            # header_cells: DATE, RFQ REF, QTN REF bilgilerini topla
            for row_idx in range(min(25, len(df))):
                for col_idx in range(min(10, len(df.columns))):
                    cell_val = str(df.iloc[row_idx, col_idx]).strip() if pd.notna(df.iloc[row_idx, col_idx]) else ''
                    if cell_val.upper().startswith('DATE') and ':' in cell_val:
                        next_col = col_idx + 1
                        val = str(df.iloc[row_idx, next_col]).strip() if next_col < len(df.columns) and pd.notna(df.iloc[row_idx, next_col]) else ''
                        header_cells.append(('DATE', val))
                        # RFQ ve QTN genelde hemen altında
                        for offset in range(1, 3):
                            r = row_idx + offset
                            if r < len(df):
                                lbl = str(df.iloc[r, col_idx]).strip() if pd.notna(df.iloc[r, col_idx]) else ''
                                v = str(df.iloc[r, next_col]).strip() if next_col < len(df.columns) and pd.notna(df.iloc[r, next_col]) else ''
                                clean_lbl = lbl.rstrip(' :').strip()
                                header_cells.append((clean_lbl, v))
                        break
                if header_cells:
                    break

            # Currency: TOTAL satırından algıla
            currency = ''
            for idx in range(len(df)):
                for col_idx in range(len(df.columns)):
                    cell_val = str(df.iloc[idx, col_idx]).strip() if pd.notna(df.iloc[idx, col_idx]) else ''
                    if 'TOTAL' in cell_val.upper() and ':' in cell_val:
                        # Yanındaki sütunda para birimi olabilir
                        for c in range(col_idx + 1, min(col_idx + 3, len(df.columns))):
                            if pd.notna(df.iloc[idx, c]):
                                total_str = str(df.iloc[idx, c])
                                for sym, code in [('€', 'EUR'), ('$', 'USD'), ('£', 'GBP'), ('₺', 'TRY')]:
                                    if sym in total_str:
                                        currency = code
                                        break
                                if currency:
                                    break
                if currency:
                    break

            header_info['currency'] = currency

            # Data satırlarını bul
            start_row = None
            for idx in range(len(df)):
                cell_val = df.iloc[idx, 0]
                if pd.notna(cell_val) and str(cell_val).strip().upper() == 'NO':
                    start_row = idx
                    break

            if start_row is None:
                return None

            data_rows = []
            for idx in range(start_row + 1, len(df)):
                row = df.iloc[idx]
                first_col_val = row.iloc[0]

                if pd.isna(first_col_val) or str(first_col_val).strip() == '':
                    # TOTAL tespiti: sadece F sütununda (index 5) "TOTAL" aranır
                    # Böylece REMARKS veya başka sütunlarda "TOTAL" geçmesi sorun yaratmaz
                    if len(row) > 5 and pd.notna(row.iloc[5]):
                        f_val = str(row.iloc[5]).strip().upper()
                        if 'TOTAL' in f_val:
                            break
                    continue

                val_str = str(first_col_val).strip()
                if val_str and val_str[0].isdigit():
                    data_rows.append(row.values.tolist())
                else:
                    continue

            return {
                'file_name': Path(file_path).name,
                'header_info': header_info,
                'header_cells': header_cells,
                'data_rows': data_rows,
            }
        except Exception:
            return None

    # ── Stiller ──────────────────────────────────────────────

    def _apply_header_style(self, ws, row_num):
        for col in range(1, 12):
            cell = ws.cell(row_num, col)
            cell.fill = self._header_fill
            cell.font = self._header_font
            cell.alignment = self._center_align
            cell.border = self._thin_border

    def _apply_data_row_style(self, ws, row_num):
        for col in range(1, 12):
            cell = ws.cell(row_num, col)
            cell.border = self._thin_border
            cell.alignment = self._data_align

    def _apply_total_style(self, ws, row_num, label, col_label=6, col_value=7):
        ws.cell(row_num, col_label).value = label
        ws.cell(row_num, col_label).font = self._bold_font
        ws.cell(row_num, col_label).alignment = self._right_align
        ws.cell(row_num, col_value).font = self._bold_font

    # ── Dosya Açma ───────────────────────────────────────────

    def open_file(self):
        if self.output_path and self.output_path.exists():
            try:
                os.startfile(str(self.output_path))
            except Exception:
                pass


def main():
    if HAS_DND:
        class DnDCTk(ctk.CTk, TkinterDnD.DnDWrapper):
            def __init__(self):
                super().__init__()
                self.TkdndVersion = TkinterDnD._require(self)
        root = DnDCTk()
    else:
        root = ctk.CTk()
    OrderSummaryMerger(root)
    root.mainloop()


if __name__ == '__main__':
    main()
